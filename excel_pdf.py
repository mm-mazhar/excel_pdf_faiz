import os
from typing import Any

import openpyxl
import pandas as pd
from tqdm import tqdm

from utils import *

# Set the option to display all rows
pd.set_option("display.max_rows", None)
# Set the option to display all rows
pd.set_option("display.max_rows", None)

# _________________________#
# ---| LOAD DATAFRAMES |---#
# -------------------------#
cert_df: pd.DataFrame = pd.read_excel(excel_file_path, sheet_name=sheet_name_1)
# print(cert_df.iloc[0:4])
# print(cert_df.columns)

data_df: pd.DataFrame = pd.read_excel(
    excel_file_path, sheet_name=sheet_name_2, header=data_df_headers
)
# Apply the function to each level in the MultiIndex
new_columns: list[tuple[str | Any, str]] = [
    (format_datetime(level1), level2) for level1, level2 in data_df.columns
]
# print(new_columns)
# Reassign the formatted columns to the DataFrame
data_df.columns = pd.MultiIndex.from_tuples(new_columns)
# print(data_df.columns)
# print(data_df.iloc[0:3])

# Ensure the output directory exists and is empty
ensure_clean_directory(os.path.dirname(output_pdfs))

# ______________________#
# ---| WRITING PDFs |---#
# ----------------------#
for i in tqdm(
    range(len(data_df)),
    colour="green",
    desc="Converting Excel Sheets to PDF files",
):

    PASSWORD: str = (
        "0" + str(data_df[data_password][i])
        if len(str(data_df[data_password][i])) == 3
        else str(data_df[data_password][i])
    )
    # print(f"PASSWORD: {PASSWORD}")

    EMPLOYEE_ID_NO: str = (
        "0" + str(data_df[data_Employee_Id_No][i])
        if len(str(data_df[data_Employee_Id_No][i])) == 3
        else str(data_df[data_Employee_Id_No][i])
    )
    # print(f"EMPLOYEE ID: {EMPLOYEE_ID_NO}")
    # print(type(EMPLOYEE_ID_NO))

    TAX: int = data_df[data_tax][i]
    # print(f"TAX: {TAX}")

    SPELL_NO: str = data_df[data_spell_no][i]
    # print(f"AMOUNT IN WORDS: {SPELL_NO}")

    NAME: str = data_df[data_name][i]
    # print(f"NAME: {NAME}")

    ADDRESS: str = data_df[data_address][i]
    # print(f"ADDRESS: {ADDRESS}")

    CNIC: str = data_df[data_cnic][i]
    # print(f"CNIC: {CNIC}")

    AMOUNT: int = data_df[data_amount][i]
    # print(f"AMOUNT: {AMOUNT}")

    DATE_YY_7: str = data_df[data_date_2023_7][i]
    # print(f"DEPOSIT DATE: {DATE_YY_7}")

    DATE_TAX_YY_7: int = data_df[data_tax_2023_7][i]
    # print(f"TAX OF MONTH JUL: {DATE_TAX_YY_7}")

    DATE_ID_YY_7: str = data_df[data_id_2023_7][i]
    # print(f"Challan / Treasury No/CPR No.: {DATE_ID_YY_7}")

    DATE_YY_8: str = data_df[data_date_2023_8][i]
    # print(f"DEPOSIT DATE: {DATE_YY_8}")

    DATE_TAX_YY_8: int = data_df[data_tax_2023_8][i]
    # print(f"TAX OF MONTH AUG: {DATE_TAX_YY_8}")

    DATE_ID_YY_8: str = data_df[data_id_2023_8][i]
    # print(f"Challan / Treasury No/CPR No.: {DATE_ID_YY_8}")

    DATE_YY_9: str = data_df[data_date_2023_9][i]
    # print(f"DEPOSIT DATE: {DATE_YY_9}")

    DATE_TAX_YY_9: int = data_df[data_tax_2023_9][i]
    # print(f"TAX OF MONTH SEPT: {DATE_TAX_YY_9}")

    DATE_ID_YY_9: str = data_df[data_id_2023_9][i]
    # print(f"Challan / Treasury No/CPR No.: {DATE_ID_YY_9}")

    DATE_YY_10: str = data_df[data_date_2023_10][i]
    # print(f"DEPOSIT DATE: {DATE_YY_10}")

    DATE_TAX_YY_10: int = data_df[data_tax_2023_10][i]
    # print(f"TAX OF MONTH OCT: {DATE_TAX_YY_10}")

    DATE_ID_YY_10: str = data_df[data_id_2023_10][i]
    # print(f"Challan / Treasury No/CPR No.: {DATE_ID_YY_10}")

    DATE_YY_11: str = data_df[data_date_2023_11][i]
    # print(f"DEPOSIT DATE: {DATE_YY_11}")

    DATE_TAX_YY_11: int = data_df[data_tax_2023_11][i]
    # print(f"TAX OF MONTH NOV: {DATE_TAX_YY_11}")

    DATE_ID_YY_11: str = data_df[data_id_2023_11][i]
    # print(f"Challan / Treasury No/CPR No.: {DATE_ID_YY_11}")

    DATE_YY_12: str = data_df[data_date_2023_12][i]
    # print(f"DEPOSIT DATE: {DATE_YY_12}")

    DATE_TAX_YY_12: int = data_df[data_tax_2023_12][i]
    # print(f"TAX OF MONTH DEC: {DATE_TAX_YY_12}")

    DATE_ID_YY_12: str = data_df[data_id_2023_12][i]
    # print(f"Challan / Treasury No/CPR No.: {DATE_ID_YY_12}")

    DATE_YY_1: str = data_df[data_date_2024_1][i]
    # print(f"DEPOSIT DATE: {DATE_YY_1}")

    DATE_TAX_YY_1: int = data_df[data_tax_2024_1][i]
    # print(f"TAX OF MONTH JAN: {DATE_TAX_YY_1}")

    DATE_ID_YY_1: str = data_df[data_id_2024_1][i]
    # print(f"Challan / Treasury No/CPR No.: {DATE_ID_YY_1}")

    DATE_YY_2: str = data_df[data_date_2024_2][i]
    # print(f"DEPOSIT DATE: {DATE_YY_2}")

    DATE_TAX_YY_2: int = data_df[data_tax_2024_2][i]
    # print(f"TAX OF MONTH FEB: {DATE_TAX_YY_2}")

    DATE_ID_YY_2: str = data_df[data_id_2024_2][i]
    # print(f"Challan / Treasury No/CPR No.: {DATE_ID_YY_2}")

    DATE_YY_3: str = data_df[data_date_2024_3][i]
    # print(f"DEPOSIT DATE: {DATE_YY_3}")

    DATE_TAX_YY_3: int = data_df[data_tax_2024_3][i]
    # print(f"TAX OF MONTH MAR: {DATE_TAX_YY_3}")

    DATE_ID_YY_3: str = data_df[data_id_2024_3][i]
    # print(f"Challan / Treasury No/CPR No.: {DATE_ID_YY_3}")

    DATE_YY_4: str = data_df[data_date_2024_4][i]
    # print(f"DEPOSIT DATE: {DATE_YY_4}")

    DATE_TAX_YY_4: int = data_df[data_tax_2024_4][i]
    # print(f"TAX OF MONTH APR: {DATE_TAX_YY_4}")

    DATE_ID_YY_4: str = data_df[data_id_2024_4][i]
    # print(f"Challan / Treasury No/CPR No.: {DATE_ID_YY_4}")

    DATE_YY_5: str = data_df[data_date_2024_5][i]
    # print(f"DEPOSIT DATE: {DATE_YY_5}")

    DATE_TAX_YY_5: int = data_df[data_tax_2024_5][i]
    # print(f"TAX OF MONTH MAY: {DATE_TAX_YY_5}")

    DATE_ID_YY_5: str = data_df[data_id_2024_5][i]
    # print(f"Challan / Treasury No/CPR No.: {DATE_ID_YY_5}")

    DATE_YY_6: str = data_df[data_date_2024_6][i]
    # print(f"DEPOSIT DATE: {DATE_YY_6}")

    DATE_TAX_YY_6: int = data_df[data_tax_2024_6][i]
    # print(f"TAX OF MONTH JUN: {DATE_TAX_YY_6}")

    DATE_ID_YY_6: str = data_df[data_id_2024_6][i]
    # print(f"Challan / Treasury No/CPR No.: {DATE_ID_YY_6}")

    # print("-" * 50)
    # print("DATA BEING SAVED TO EXCEL FILE AND CONVERTING SHEET TO PDF FILE")
    # print("-" * 50)

    # # Load the workbook and the sheet
    wb: openpyxl.Workbook = openpyxl.load_workbook(excel_file_path)
    ws: Any = wb[sheet_name_1]

    # Update the cell With the data acquired
    ws[cert_EMPLOYEE_ID_NO] = EMPLOYEE_ID_NO  # Replace with your EMPLOYEE_ID_NO
    ws[cert_TAX] = TAX  # Replace with your TAX
    ws[cert_SPELL_NO] = SPELL_NO  # Replace with your SPELL_NO
    ws[cert_NAME] = NAME  # Replace with your NAME
    ws[cert_ADDRESS] = ADDRESS  # Replace with your ADDRESS
    ws[cert_CNIC] = CNIC  # Replace with your CNIC
    ws[cert_AMOUNT] = AMOUNT  # Replace with your AMOUNT

    ws[cert_DATE_YY_7] = DATE_YY_7  # Replace with your DATE_YY_7
    ws[cert_DATE_TAX_YY_7] = DATE_TAX_YY_7  # Replace with your DATE_TAX_YY_7
    ws[cert_DATE_ID_YY_7] = DATE_ID_YY_7  # Replace with your DATE_ID_YY_7

    ws[cert_DATE_YY_8] = DATE_YY_8  # Replace with your DATE_YY_8
    ws[cert_DATE_TAX_YY_8] = DATE_TAX_YY_8  # Replace with your DATE_TAX_YY_8
    ws[cert_DATE_ID_YY_8] = DATE_ID_YY_8  # Replace with your DATE_ID_YY_8

    ws[cert_DATE_YY_9] = DATE_YY_9  # Replace with your DATE_YY_9
    ws[cert_DATE_TAX_YY_9] = DATE_TAX_YY_9  # Replace with your DATE_TAX_YY_9
    ws[cert_DATE_ID_YY_9] = DATE_ID_YY_9  # Replace with your DATE_ID_YY_9

    ws[cert_DATE_YY_10] = DATE_YY_10  # Replace with your DATE_YY_10
    ws[cert_DATE_TAX_YY_10] = DATE_TAX_YY_10  # Replace with your DATE_TAX_YY_10
    ws[cert_DATE_ID_YY_10] = DATE_ID_YY_10  # Replace with your DATE_ID_YY_10

    ws[cert_DATE_YY_11] = DATE_YY_11  # Replace with your DATE_YY_11
    ws[cert_DATE_TAX_YY_11] = DATE_TAX_YY_11  # Replace with your DATE_TAX_YY_11
    ws[cert_DATE_ID_YY_11] = DATE_ID_YY_11  # Replace with your DATE_ID_YY_11

    ws[cert_DATE_YY_12] = DATE_YY_12  # Replace with your DATE_YY_12
    ws[cert_DATE_TAX_YY_12] = DATE_TAX_YY_12  # Replace with your DATE_TAX_YY_12
    ws[cert_DATE_ID_YY_12] = DATE_ID_YY_12  # Replace with your DATE_ID_YY_12
    ws[cert_DATE_YY_1] = DATE_YY_1  # Replace with your DATE_YY_1
    ws[cert_DATE_TAX_YY_1] = DATE_TAX_YY_1  # Replace with your DATE_TAX_YY_1
    ws[cert_DATE_ID_YY_1] = DATE_ID_YY_1  # Replace with your DATE_ID_YY_1
    ws[cert_DATE_YY_2] = DATE_YY_2  # Replace with your DATE_YY_2
    ws[cert_DATE_TAX_YY_2] = DATE_TAX_YY_2  # Replace with your DATE_TAX_YY_2
    ws[cert_DATE_ID_YY_2] = DATE_ID_YY_2  # Replace with your DATE_ID_YY_2
    ws[cert_DATE_YY_3] = DATE_YY_3  # Replace with your DATE_YY_3
    ws[cert_DATE_TAX_YY_3] = DATE_TAX_YY_3  # Replace with your DATE_TAX_YY_3
    ws[cert_DATE_ID_YY_3] = DATE_ID_YY_3  # Replace with your DATE_ID_YY_3
    ws[cert_DATE_YY_4] = DATE_YY_4  # Replace with your DATE_YY_4
    ws[cert_DATE_TAX_YY_4] = DATE_TAX_YY_4  # Replace with your DATE_TAX_YY_4
    ws[cert_DATE_ID_YY_4] = DATE_ID_YY_4  # Replace with your DATE_ID_YY_4
    ws[cert_DATE_YY_5] = DATE_YY_5  # Replace with your DATE_YY_5
    ws[cert_DATE_TAX_YY_5] = DATE_TAX_YY_5  # Replace with your DATE_TAX_YY_5
    ws[cert_DATE_ID_YY_5] = DATE_ID_YY_5  # Replace with your DATE_ID_YY_5

    # Save the updated Excel file
    wb.save(excel_file_path)
    # After saving, if you're done with the workbook, you can delete it
    del wb  # This will release the workbook object
    # print("Data Saved Successfully on Excel Sheet and Converting Sheet to PDF")

    pdf_file_name: str = NAME.replace(" ", "_")
    pdf_file_name = os.path.join(output_pdfs, f"{pdf_file_name}.pdf")

    # print(f"File: {pdf_file_name}")

    convert_excel_to_pdf_with_password(
        excel_file_path, sheet_name_1, pdf_file_name, PASSWORD
    )

    # print(f"SAVED: {pdf_file_name}")

    # print("*" * 120)

print(f"PDF FILES SAVED To: {output_pdfs}")
